import os
import math
import json 
import backend
import openpyxl
import xlsxwriter
from utils import *
from analyst import analytics
from openpyxl.styles import PatternFill
from PIL import Image,  ExifTags
from io import BytesIO
from api import TBA

config_file = json.load(open('assets/config.json'))
colors = config_file["colors"]

def contribution_csv(color=True):
    teams = backend.allTeams()
    csv_list = []
    header = [
            "Team",
            "Auton Low",
            "Auton Mid",
            "Auton High",
            "Teleop Low",
            "Teleop Mid",
            "Teleop High",
            "Low Cones Teleop",
            "Mid Cones Teleop",
            "High Cones Teleop",
            "Low Cubes Teleop",
            "Mid Cubes Teleop",
            "High Cubes Teleop",
            "Cones Teleop",
            "Cubes Teleop",
            "Pieces Teleop",
            "Auton Dock %",
            "Auton Engage %",
            "Teleop Dock %",
            "Teleop Engage %",
            "Auton Pts",
            "Teleop Pts",
            "Endgame Pts",
            "Total Pts",
        ]
    for team in teams:
        analyst = analytics(team)
        docked_auto, engaged_auto, len_auton = analyst.format_charge("auton")
        docked_teleop, engaged_teleop, len_teleop = analyst.format_charge("teleop")
        team_data = [
            team,
            get_average(analyst.get_list_cargo_general("L", "auton")),
            get_average(analyst.get_list_cargo_general("M", "auton")),
            get_average(analyst.get_list_cargo_general("H", "auton")),
            get_average(analyst.get_list_cargo_general("L", "teleop")),
            get_average(analyst.get_list_cargo_general("M", "teleop")),
            get_average(analyst.get_list_cargo_general("H", "teleop")),
            get_average(analyst.get_list_cargo_specific("L", "teleop", "cone")),
            get_average(analyst.get_list_cargo_specific("M", "teleop", "cone")),
            get_average(analyst.get_list_cargo_specific("H", "teleop", "cone")),
            get_average(analyst.get_list_cargo_specific("L", "teleop", "cube")),
            get_average(analyst.get_list_cargo_specific("M", "teleop", "cube")),
            get_average(analyst.get_list_cargo_specific("H", "teleop", "cube")),
            get_average(analyst.get_list_cargo_specific_ignore_level("cone", "teleop")), 
            get_average(analyst.get_list_cargo_specific_ignore_level("cube", "teleop")), 
            get_average(analyst.get_piece_progression()),
            round(docked_auto/len_auton, 2),
            round(engaged_auto/len_auton, 2),
            round(docked_teleop/len_teleop, 2),
            round(engaged_teleop/len_teleop, 2),
            analyst.get_point_average("auton"),
            analyst.get_point_average("teleop"),
            analyst.get_point_average("endgame"),
            analyst.get_point_average("total")
        ]
        csv_list.append(team_data)
    csv_list = sorted(csv_list, key=lambda x:x[-1], reverse=True)
    csv_list.insert(0, header) 
    save(csv_list, "averages")
    if color:
        colorCode("averages")

def colorCode(filename):
    assert os.path.isfile(filename+".xlsx")
    doc = openpyxl.load_workbook(filename+".xlsx")
    spreadsheet = doc.active
    for column in range(2, spreadsheet.max_column+1):
        column_data = []
        for row in range(2, spreadsheet.max_row+1):
            column_data.append(spreadsheet.cell(row=row, column=column).value)
        column_data = sorted(column_data, reverse=True)
        for row in range(2, spreadsheet.max_row+1):
            cell = spreadsheet.cell(row=row, column=column)
            index = math.floor((len(colors)-1)*column_data.index(cell.value)/(len(column_data)-1))
            if cell.value==0 and index!=7:
                index==6
            color = colors[index]
            cell.fill = PatternFill(fgColor=color, bgColor=color, fill_type='solid')
    doc.save(filename+".xlsx")

def fill_sheet(sheet, data, start_row, start_column):
    for row in range(0, len(data)):
        for column in range(0, len(data[0])):
            sheet.write(row+start_row, column+start_column, data[row][column])
            
def generate_team_by_team():
    workbook = xlsxwriter.Workbook('team_by_team.xlsx')
    TBA().download_pictures()
    for team in backend.allTeams():
        worksheet = workbook.add_worksheet(name=team)
        analytic = analytics(team)
        progression = analytic.csv_progression()
        fill_sheet(worksheet, progression, 0, 0)
        starting_index = len(progression)+2
        fill_sheet(worksheet, analytic.get_csv_summary(), starting_index, 6)
        try:
            image_path = 'pictures/'+team+'.jpeg'
            image = Image.open(image_path)
            try:
                for orientation in ExifTags.TAGS.keys():
                    if ExifTags.TAGS[orientation]=='Orientation':
                        break
                exif=dict(image._getexif().items())
                if exif[orientation] == 3:
                    image=image.rotate(180, expand=True)
                elif exif[orientation] == 6:
                    image=image.rotate(270, expand=True)
                elif exif[orientation] == 8:
                    image=image.rotate(90, expand=True)
            except (AttributeError, KeyError, IndexError):
                pass
            aspect_ratio = float(image.size[1]) / float(image.size[0])
            resized_image = image.resize((300, int(300 * aspect_ratio)))
            image_buffer = BytesIO()
            resized_image.save(image_buffer, format='JPEG')
            worksheet.insert_image('A'+str(starting_index), '', {'image_data': image_buffer})
        except:
            print("NO IMAGE FOR", team)
    workbook.close()

def create_spreadsheets():
    contribution_csv()
    generate_team_by_team()