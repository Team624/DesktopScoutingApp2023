import openpyxl
import math
import json 
import os
from openpyxl.styles import PatternFill
import backend
from html_editor import generate_html_loop

config_file = json.load(open('assets/config.json'))
colors = config_file["colors"]

def colorCode(filename):
    assert os.path.isfile(filename+".xlsx")
    doc = openpyxl.load_workbook(filename+".xlsx")
    spreadsheet = doc.active
    for column in range(2, spreadsheet.max_column+1):
        column_data = []
        for row in range(2, spreadsheet.max_row+1):
            column_data.append(spreadsheet.cell(row=row, column=column).value)
        column_data = sorted(column_data, reverse=True)
        for row in range(2, spreadsheet.max_row+1):
            cell = spreadsheet.cell(row=row, column=column)
            index = math.floor((len(colors)-1)*column_data.index(cell.value)/(len(column_data)-1))
            if cell.value==0 and index!=7:
                index==6
            color = colors[index]
            cell.fill = PatternFill(fgColor=color, bgColor=color, fill_type='solid')
    doc.save(filename+".xlsx")

def generate_autons(team):
    data = backend.search(team)
    generate_html_loop(data)

def get_all_autons():
    for team in backend.allTeams():
        generate_autons(team)